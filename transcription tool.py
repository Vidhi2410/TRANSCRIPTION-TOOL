import os
import re
import wave
import json
import subprocess
from vosk import Model, KaldiRecognizer
from openpyxl import load_workbook, Workbook
from PySide6 import QtWidgets, QtGui, QtCore
import sys
from pathlib import Path

# --- Paths Default ---
model_path = r"C:\Users\Amar\Desktop\vosk-model-hi-0.22"
template_path = r"C:\Users\Amar\Desktop\test_format\format.xlsx"

# --- Convert Audio to WAV ---
def convert_to_wav(input_file, output_wav="temp.wav"):
    cmd = ["ffmpeg", "-y", "-i", input_file, "-ar", "16000", "-ac", "1", output_wav]
    subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return output_wav

# --- Transcribe Audio ---
def transcribe(file_path):
    wav = convert_to_wav(file_path)
    wf = wave.open(wav, "rb")
    model = Model(model_path)
    rec = KaldiRecognizer(model, wf.getframerate())
    text = ""
    while True:
        data = wf.readframes(4000)
        if not data:
            break
        if rec.AcceptWaveform(data):
            text += json.loads(rec.Result()).get("text", "") + " "
    text += json.loads(rec.FinalResult()).get("text", "")
    return text.strip()

# --- Extract Names & Relations ---
def process_text(text):
    results = []

    # suffixes जो नाम में ही रहने चाहिए
    name_suffixes = ["जी", "बाई", "बाइ", "वाई", "देवी", "कुमारी", "लाल", "प्रसाद"]

    pattern = re.compile(
        r"([a-zA-Z\u0900-\u097F]+(?:\s*(?:" + "|".join(name_suffixes) + r"))*)"
        r"(?:\s+(सिंह|पटेल))?"
        r"(?:\s+(.*?(?:का|की|के)\s+[a-zA-Z\u0900-\u097F]+))?"
    )

    matches = pattern.findall(text)

    data_position = 0
    start_keywords = ["पुत्र", "पुत्री", "बेटा", "बेटी", "पत्नी", "पति", "धनी", "पिता", "दादा"]

    for fname, surname, relation in matches:
        fname = fname.strip()
        surname = (surname or "").strip()
        relation = (relation or "").strip()

        clean_relation = ""
        extra_name = ""

        if relation:
            parts = relation.split()
            for i, p in enumerate(parts):
                if p in start_keywords:
                    clean_relation = p
                    if i + 1 < len(parts):
                        extra_name = " ".join(parts[i + 1:])
                    break

        if data_position == 0:
            data_position = 1
        elif clean_relation:
            data_position += 1

        results.append((fname, surname, clean_relation, data_position))

        if extra_name:
            results.append((extra_name, "", "", data_position))

    return results

# --- Save to Excel (folder based) ---
def save_to_excel(data, folder_path):
    folder_path = Path(folder_path)
    folder_path.mkdir(parents=True, exist_ok=True)

    output_file = folder_path / "family_output.xlsx"

    if not output_file.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["Given Name", "Surname", "Relation", "Data Position"])
        wb.save(output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    for fname, surname, relation, data_pos in data:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=fname)
        ws.cell(row=row, column=2, value=surname)
        ws.cell(row=row, column=3, value=relation)
        ws.cell(row=row, column=4, value=data_pos)

    wb.save(output_file)
    return f"✅ Data saved to {output_file}"

# --- GUI Class ---
class FamilyApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Family Relation Extractor")
        self.setGeometry(200, 100, 900, 650)

        layout = QtWidgets.QVBoxLayout()

        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        self.loadBtn = QtWidgets.QPushButton("Load Audio File")
        self.loadBtn.clicked.connect(self.load_audio)
        self.selectFolderBtn = QtWidgets.QPushButton("Select Output Folder")
        self.selectFolderBtn.clicked.connect(self.select_folder)
        btn_layout.addWidget(self.loadBtn)
        btn_layout.addWidget(self.selectFolderBtn)
        layout.addLayout(btn_layout)

        self.transcriptBox = QtWidgets.QTextEdit()
        self.transcriptBox.setPlaceholderText("Transcript will appear here...")
        layout.addWidget(self.transcriptBox)

        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Given Name", "Surname", "Relation", "Data Position"])
        layout.addWidget(self.table)

        self.saveBtn = QtWidgets.QPushButton("Save to Selected Folder")
        self.saveBtn.clicked.connect(self.save_data)
        layout.addWidget(self.saveBtn)

        self.setLayout(layout)
        self.data = []
        self.output_folder = Path.home() / "Desktop"  # default folder

    def select_folder(self):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_folder = Path(folder)

    def load_audio(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Audio", "", "Audio Files (*.mp3 *.wav *.aac)")
        if file_path:
            transcript = transcribe(file_path)
            self.transcriptBox.setText(transcript)

            self.data = process_text(transcript)

            self.table.setRowCount(len(self.data))
            for i, (fname, surname, relation, pos) in enumerate(self.data):
                self.table.setItem(i, 0, QtWidgets.QTableWidgetItem(fname))
                self.table.setItem(i, 1, QtWidgets.QTableWidgetItem(surname))
                self.table.setItem(i, 2, QtWidgets.QTableWidgetItem(relation))
                self.table.setItem(i, 3, QtWidgets.QTableWidgetItem(str(pos)))

    def save_data(self):
        if not self.data:
            QtWidgets.QMessageBox.warning(self, "Error", "No data to save!")
            return
        msg = save_to_excel(self.data, self.output_folder)
        QtWidgets.QMessageBox.information(self, "Saved", msg)

# --- Run ---
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = FamilyApp()
    window.show()
    sys.exit(app.exec_())
